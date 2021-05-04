# home page class containing all the objects for our website home to be used in test script
import pytest
from selenium.webdriver.common.by import By
import openpyxl

class HomePage:

    def __init__(self, driver):   # the constructor method will be initialized at object creation and diver will be called from test script
        self.driver = driver

    shop = (By.LINK_TEXT,"Shop") # shop object created by using objcet syntax tuple containing locator and element
    name = (By.NAME,"name")
    email = (By.NAME, "email")
    password = (By.ID,"exampleInputPassword1")
    alert = (By.XPATH,"//div[@class='alert alert-success alert-dismissible']")
    submit = (By.XPATH,"//input[@type='submit']")



    def shopItems(self):
        return self.driver.find_element(*HomePage.shop)  # to deconstruct tuple passed as class variable use * at beginning
        #### the above step it woull represent the self.driver.find_element_by_link_text("Shop") in non object page test command

    def getName(self):
        return self.driver.find_element(*HomePage.name)
    def getEmail(self):
        return self.driver.find_element(*HomePage.email)
    def getPassword(self):
        return self.driver.find_element(*HomePage.password)
    def alert(self):
        return self.driver.find_element(*HomePage.alert)
    def sumbit(self):
        return self.driver.find_element(*HomePage.submit)



## below utility is to use data from exel spreadsheet into submitForm test case #############################

    @staticmethod  # by declaring method as static you can call it without initilizing an object for the Class
    def getTestDataExl(test_case_name):
        book = openpyxl.load_workbook(
            "C:\\Users\\u6017127\\PycharmProjects\\SeleniumPython_EndToEnd_Framework\\HomePageData.xlsx")
        sheet = book.active  # make the objet active to be used
        Dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # we need to send the data wrapped in list to be executed in test with pytest fixture

        return [Dict]  # now it is eturned wrapped into [] as a list to be used in pythest fixture